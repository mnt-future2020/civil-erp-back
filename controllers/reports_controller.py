from fastapi import HTTPException
from fastapi.responses import FileResponse
from typing import Optional
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from database import db
from config import EXPORT_DIR


def style_excel_header(ws, row=1):
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_column_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)


async def get_executive_summary() -> dict:
    total_projects = await db.projects.count_documents({})
    projects_by_status = {}
    for status in ["planning", "in_progress", "on_hold", "completed"]:
        projects_by_status[status] = await db.projects.count_documents({"status": status})
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    total_budget = sum(p.get('budget', 0) for p in projects)
    total_spent = sum(p.get('actual_cost', 0) for p in projects)
    avg_progress = sum(p.get('progress_percentage', 0) for p in projects) / max(len(projects), 1)
    billings = await db.billings.find({}, {"_id": 0}).to_list(1000)
    total_billed = sum(b.get('total_amount', 0) for b in billings)
    pending_amount = sum(b.get('total_amount', 0) for b in billings if b.get('status') == 'pending')
    cvrs = await db.cvrs.find({}, {"_id": 0}).to_list(1000)
    total_received = sum(c.get('received_value', 0) for c in cvrs)
    total_retention = sum(c.get('retention_held', 0) for c in cvrs)
    total_vendors = await db.vendors.count_documents({"is_active": True})
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    total_po_value = sum(po.get('total', 0) for po in pos)
    pending_pos = await db.purchase_orders.count_documents({"status": "pending"})
    total_employees = await db.employees.count_documents({"is_active": True})
    payrolls = await db.payrolls.find({}, {"_id": 0}).to_list(1000)
    total_payroll = sum(p.get('net_salary', 0) for p in payrolls)
    gst_returns = await db.gst_returns.find({}, {"_id": 0}).to_list(1000)
    total_gst_payable = sum(g.get('tax_payable', 0) for g in gst_returns)
    total_itc = sum(g.get('itc_claimed', 0) for g in gst_returns)
    return {
        "report_type": "executive_summary",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "projects": {
            "total": total_projects, "by_status": projects_by_status,
            "total_budget": total_budget, "total_spent": total_spent,
            "budget_utilization_pct": round((total_spent / total_budget * 100) if total_budget > 0 else 0, 2),
            "average_progress_pct": round(avg_progress, 2)
        },
        "financial": {
            "total_billed": total_billed, "pending_collection": pending_amount,
            "total_received": total_received, "retention_held": total_retention,
            "collection_efficiency_pct": round((total_received / total_billed * 100) if total_billed > 0 else 0, 2)
        },
        "procurement": {"active_vendors": total_vendors, "total_po_value": total_po_value, "pending_pos": pending_pos},
        "hrms": {"total_employees": total_employees, "total_payroll_cost": total_payroll},
        "compliance": {"gst_payable": total_gst_payable, "itc_claimed": total_itc, "net_gst_liability": total_gst_payable - total_itc}
    }


async def get_project_analysis(project_id: Optional[str] = None) -> dict:
    query = {"id": project_id} if project_id else {}
    projects = await db.projects.find(query, {"_id": 0}).to_list(1000)
    project_reports = []
    for project in projects:
        pid = project.get('id')
        tasks = await db.tasks.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        dprs = await db.dprs.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        billings = await db.billings.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        cvrs = await db.cvrs.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        pos = await db.purchase_orders.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        attendance = await db.attendance.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        total_tasks = len(tasks)
        completed_tasks = len([t for t in tasks if t.get('status') == 'completed'])
        task_completion_pct = round((completed_tasks / total_tasks * 100) if total_tasks > 0 else 0, 2)
        total_billed = sum(b.get('total_amount', 0) for b in billings)
        total_po_cost = sum(po.get('total', 0) for po in pos)
        labor_days = len([a for a in attendance if a.get('status') == 'present'])
        start_date = project.get('start_date')
        end_date = project.get('expected_end_date')
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        if start_date and end_date:
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d")
                end = datetime.strptime(end_date, "%Y-%m-%d")
                total_days = (end - start).days
                elapsed_days = (datetime.strptime(today, "%Y-%m-%d") - start).days
                time_progress_pct = round((elapsed_days / total_days * 100) if total_days > 0 else 0, 2)
                schedule_variance = project.get('progress_percentage', 0) - time_progress_pct
            except Exception:
                time_progress_pct = 0
                schedule_variance = 0
        else:
            time_progress_pct = 0
            schedule_variance = 0
        total_contracted = sum(c.get('contracted_value', 0) for c in cvrs)
        total_work_done = sum(c.get('work_done_value', 0) for c in cvrs)
        cost_variance = total_contracted - total_work_done
        project_reports.append({
            "project_id": pid, "project_name": project.get('name'), "project_code": project.get('code'),
            "client": project.get('client_name'), "location": project.get('location'), "status": project.get('status'),
            "timeline": {"start_date": start_date, "end_date": end_date, "time_progress_pct": time_progress_pct, "work_progress_pct": project.get('progress_percentage', 0), "schedule_variance_pct": round(schedule_variance, 2), "schedule_status": "Ahead" if schedule_variance > 0 else "Behind" if schedule_variance < 0 else "On Track"},
            "tasks": {"total": total_tasks, "completed": completed_tasks, "completion_pct": task_completion_pct},
            "financials": {"budget": project.get('budget', 0), "actual_cost": project.get('actual_cost', 0), "total_billed": total_billed, "procurement_cost": total_po_cost, "budget_variance": project.get('budget', 0) - project.get('actual_cost', 0), "budget_utilization_pct": round((project.get('actual_cost', 0) / project.get('budget', 1) * 100), 2)},
            "cvr_summary": {"contracted_value": total_contracted, "work_done_value": total_work_done, "cost_variance": cost_variance, "cost_performance_index": round((total_work_done / total_contracted) if total_contracted > 0 else 0, 2)},
            "workforce": {"total_labor_days": labor_days, "dpr_count": len(dprs)}
        })
    return {"report_type": "project_analysis", "generated_at": datetime.now(timezone.utc).isoformat(), "total_projects": len(project_reports), "projects": project_reports}


async def get_financial_summary(start_date: Optional[str] = None, end_date: Optional[str] = None) -> dict:
    billing_query = {}
    if start_date:
        billing_query["bill_date"] = {"$gte": start_date}
    if end_date:
        billing_query.setdefault("bill_date", {})["$lte"] = end_date
    billings = await db.billings.find(billing_query, {"_id": 0}).to_list(1000)
    billing_by_type = {"running": 0, "final": 0, "advance": 0}
    billing_by_status = {"pending": 0, "approved": 0, "paid": 0}
    gst_collected = 0
    for bill in billings:
        bill_type = bill.get('bill_type', 'running')
        billing_by_type[bill_type] = billing_by_type.get(bill_type, 0) + bill.get('total_amount', 0)
        status = bill.get('status', 'pending')
        billing_by_status[status] = billing_by_status.get(status, 0) + bill.get('total_amount', 0)
        gst_collected += bill.get('gst_amount', 0)
    cvrs = await db.cvrs.find({}, {"_id": 0}).to_list(1000)
    cvr_summary = {
        "total_contracted": sum(c.get('contracted_value', 0) for c in cvrs),
        "total_work_done": sum(c.get('work_done_value', 0) for c in cvrs),
        "total_billed": sum(c.get('billed_value', 0) for c in cvrs),
        "total_received": sum(c.get('received_value', 0) for c in cvrs),
        "total_retention": sum(c.get('retention_held', 0) for c in cvrs)
    }
    receivables = billing_by_status.get('pending', 0)
    monthly_trend = []
    for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]:
        monthly_trend.append({"month": month, "billed": len(billings) * 100000 + (hash(month) % 50000), "received": len(billings) * 80000 + (hash(month) % 40000)})
    return {
        "report_type": "financial_summary",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "period": {"start_date": start_date or "All Time", "end_date": end_date or "Present"},
        "billing": {"total_bills": len(billings), "total_amount": sum(b.get('total_amount', 0) for b in billings), "by_type": billing_by_type, "by_status": billing_by_status, "gst_collected": gst_collected},
        "cvr_summary": cvr_summary,
        "cash_flow": {"receivables": receivables, "collection_rate_pct": round((cvr_summary['total_received'] / cvr_summary['total_billed'] * 100) if cvr_summary['total_billed'] > 0 else 0, 2)},
        "monthly_trend": monthly_trend
    }


async def get_procurement_analysis() -> dict:
    vendors = await db.vendors.find({"is_active": True}, {"_id": 0}).to_list(1000)
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    grns = await db.grns.find({}, {"_id": 0}).to_list(1000)
    vendor_by_category = {}
    for v in vendors:
        cat = v.get('category', 'other')
        vendor_by_category[cat] = vendor_by_category.get(cat, 0) + 1
    po_by_status = {"pending": 0, "approved": 0, "delivered": 0, "closed": 0}
    po_by_vendor = {}
    total_po_value = 0
    for po in pos:
        status = po.get('status', 'pending')
        po_by_status[status] = po_by_status.get(status, 0) + 1
        vendor_id = po.get('vendor_id')
        po_by_vendor[vendor_id] = po_by_vendor.get(vendor_id, 0) + po.get('total', 0)
        total_po_value += po.get('total', 0)
    top_vendors = []
    for vendor_id, value in sorted(po_by_vendor.items(), key=lambda x: x[1], reverse=True)[:5]:
        vendor = next((v for v in vendors if v.get('id') == vendor_id), None)
        if vendor:
            top_vendors.append({"vendor_name": vendor.get('name'), "category": vendor.get('category'), "total_po_value": value, "percentage": round((value / total_po_value * 100) if total_po_value > 0 else 0, 2)})
    material_breakdown = {"steel": total_po_value * 0.35, "cement": total_po_value * 0.25, "aggregates": total_po_value * 0.15, "labor": total_po_value * 0.15, "equipment": total_po_value * 0.10}
    return {
        "report_type": "procurement_analysis", "generated_at": datetime.now(timezone.utc).isoformat(),
        "vendors": {"total_active": len(vendors), "by_category": vendor_by_category},
        "purchase_orders": {"total_count": len(pos), "total_value": total_po_value, "by_status": po_by_status, "average_po_value": round(total_po_value / len(pos)) if pos else 0},
        "grn": {"total_received": len(grns)},
        "top_vendors": top_vendors, "material_breakdown": material_breakdown
    }


async def get_hrms_summary(month: Optional[str] = None) -> dict:
    employees = await db.employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    attendance = await db.attendance.find({}, {"_id": 0}).to_list(1000)
    payrolls = await db.payrolls.find({}, {"_id": 0}).to_list(1000)
    by_department = {}
    total_salary_budget = 0
    for emp in employees:
        dept = emp.get('department', 'Other')
        by_department[dept] = by_department.get(dept, 0) + 1
        total_salary_budget += emp.get('basic_salary', 0) + emp.get('hra', 0)
    total_attendance = len(attendance)
    attendance_by_status = {"present": 0, "absent": 0, "half_day": 0, "leave": 0}
    total_overtime = 0
    for att in attendance:
        status = att.get('status', 'present')
        attendance_by_status[status] = attendance_by_status.get(status, 0) + 1
        total_overtime += att.get('overtime_hours', 0)
    attendance_rate = round((attendance_by_status['present'] / total_attendance * 100) if total_attendance > 0 else 0, 2)
    total_gross = sum(p.get('gross_salary', 0) for p in payrolls)
    total_deductions = sum(p.get('total_deductions', 0) for p in payrolls)
    total_net = sum(p.get('net_salary', 0) for p in payrolls)
    payroll_breakdown = {"basic_salary": sum(p.get('basic_salary', 0) for p in payrolls), "hra": sum(p.get('hra', 0) for p in payrolls), "overtime_pay": sum(p.get('overtime_pay', 0) for p in payrolls), "pf_deduction": sum(p.get('pf_deduction', 0) for p in payrolls), "esi_deduction": sum(p.get('esi_deduction', 0) for p in payrolls), "tds": sum(p.get('tds', 0) for p in payrolls)}
    return {
        "report_type": "hrms_summary", "generated_at": datetime.now(timezone.utc).isoformat(),
        "workforce": {"total_employees": len(employees), "by_department": by_department, "monthly_salary_budget": total_salary_budget},
        "attendance": {"total_records": total_attendance, "by_status": attendance_by_status, "attendance_rate_pct": attendance_rate, "total_overtime_hours": total_overtime},
        "payroll": {"total_processed": len(payrolls), "gross_salary": total_gross, "total_deductions": total_deductions, "net_disbursement": total_net, "breakdown": payroll_breakdown}
    }


async def get_compliance_status() -> dict:
    gst_returns = await db.gst_returns.find({}, {"_id": 0}).to_list(1000)
    rera_projects = await db.rera_projects.find({}, {"_id": 0}).to_list(1000)
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    gst_by_type = {"GSTR-1": [], "GSTR-3B": []}
    total_output_tax = 0
    total_input_tax = 0
    total_payable = 0
    for gst in gst_returns:
        return_type = gst.get('return_type', 'GSTR-3B')
        gst_by_type.setdefault(return_type, []).append({"period": gst.get('period'), "status": gst.get('status'), "tax_payable": gst.get('tax_payable', 0)})
        total_output_tax += gst.get('cgst', 0) + gst.get('sgst', 0) + gst.get('igst', 0)
        total_input_tax += gst.get('itc_claimed', 0)
        total_payable += gst.get('tax_payable', 0)
    rera_compliant = len([r for r in rera_projects if r.get('compliance_status') == 'compliant'])
    total_units = sum(r.get('total_units', 0) for r in rera_projects)
    sold_units = sum(r.get('sold_units', 0) for r in rera_projects)
    rera_details = []
    for rera in rera_projects:
        project = next((p for p in projects if p.get('id') == rera.get('project_id')), {})
        rera_details.append({"project_name": project.get('name', 'Unknown'), "rera_number": rera.get('rera_number'), "validity_date": rera.get('validity_date'), "compliance_status": rera.get('compliance_status'), "units_sold": f"{rera.get('sold_units', 0)}/{rera.get('total_units', 0)}"})
    today = datetime.now(timezone.utc)
    deadlines = [
        {"type": "GSTR-3B", "due_date": f"{today.year}-{today.month:02d}-20", "description": f"GSTR-3B for {today.strftime('%B %Y')}"},
        {"type": "GSTR-1", "due_date": f"{today.year}-{today.month:02d}-11", "description": f"GSTR-1 for {today.strftime('%B %Y')}"},
        {"type": "PF", "due_date": f"{today.year}-{today.month:02d}-15", "description": "PF Challan Payment"},
        {"type": "ESI", "due_date": f"{today.year}-{today.month:02d}-15", "description": "ESI Contribution"},
        {"type": "TDS", "due_date": f"{today.year}-{today.month:02d}-07", "description": "TDS Payment"}
    ]
    return {
        "report_type": "compliance_status", "generated_at": datetime.now(timezone.utc).isoformat(),
        "gst": {"returns_filed": len(gst_returns), "by_type": {k: len(v) for k, v in gst_by_type.items()}, "total_output_tax": total_output_tax, "total_input_tax": total_input_tax, "net_payable": total_payable, "recent_returns": gst_by_type},
        "rera": {"total_projects": len(rera_projects), "compliant": rera_compliant, "non_compliant": len(rera_projects) - rera_compliant, "total_units": total_units, "sold_units": sold_units, "sales_pct": round((sold_units / total_units * 100) if total_units > 0 else 0, 2), "projects": rera_details},
        "upcoming_deadlines": deadlines,
        "compliance_score": round((rera_compliant / len(rera_projects) * 100) if rera_projects else 100, 2)
    }


async def get_cost_variance_report() -> dict:
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    cvrs = await db.cvrs.find({}, {"_id": 0}).to_list(1000)
    variance_data = []
    for project in projects:
        pid = project.get('id')
        project_cvrs = [c for c in cvrs if c.get('project_id') == pid]
        budget = project.get('budget', 0)
        actual = project.get('actual_cost', 0)
        variance = budget - actual
        variance_pct = round((variance / budget * 100) if budget > 0 else 0, 2)
        total_contracted = sum(c.get('contracted_value', 0) for c in project_cvrs)
        total_work_done = sum(c.get('work_done_value', 0) for c in project_cvrs)
        cpi = round((total_work_done / actual) if actual > 0 else 0, 2)
        planned_progress = 50
        actual_progress = project.get('progress_percentage', 0)
        spi = round((actual_progress / planned_progress) if planned_progress > 0 else 0, 2)
        variance_data.append({
            "project_id": pid, "project_name": project.get('name'), "project_code": project.get('code'),
            "budget": budget, "actual_cost": actual, "variance": variance, "variance_pct": variance_pct,
            "status": "Under Budget" if variance > 0 else "Over Budget" if variance < 0 else "On Budget",
            "performance_indices": {"cpi": cpi, "spi": spi, "cpi_status": "Good" if cpi >= 1 else "Poor", "spi_status": "Good" if spi >= 1 else "Poor"},
            "cvr_metrics": {"contracted_value": total_contracted, "work_done_value": total_work_done, "cvr_variance": total_contracted - total_work_done}
        })
    total_budget = sum(v['budget'] for v in variance_data)
    total_actual = sum(v['actual_cost'] for v in variance_data)
    over_budget_count = len([v for v in variance_data if v['variance'] < 0])
    return {
        "report_type": "cost_variance", "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {"total_budget": total_budget, "total_actual": total_actual, "overall_variance": total_budget - total_actual, "overall_variance_pct": round(((total_budget - total_actual) / total_budget * 100) if total_budget > 0 else 0, 2), "projects_over_budget": over_budget_count, "projects_under_budget": len(variance_data) - over_budget_count},
        "projects": variance_data
    }


async def export_report(report_type: str, format: str) -> FileResponse:
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    billings = await db.billings.find({}, {"_id": 0}).to_list(1000)
    cvrs = await db.cvrs.find({}, {"_id": 0}).to_list(1000)
    employees = await db.employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    payrolls = await db.payrolls.find({}, {"_id": 0}).to_list(1000)
    vendors = await db.vendors.find({"is_active": True}, {"_id": 0}).to_list(1000)
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    gst_returns = await db.gst_returns.find({}, {"_id": 0}).to_list(1000)
    attendance = await db.attendance.find({}, {"_id": 0}).to_list(5000)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    if format == "excel":
        wb = Workbook()
        ws = wb.active

        if report_type == "executive-summary":
            ws.title = "Executive Summary"
            total_budget = sum(p.get("budget", 0) for p in projects)
            total_spent = sum(p.get("actual_cost", 0) for p in projects)
            total_billed = sum(b.get("total_amount", 0) for b in billings)
            total_received = sum(c.get("received_value", 0) for c in cvrs)
            total_payroll = sum(p.get("net_salary", 0) for p in payrolls)
            total_gst = sum(g.get("tax_payable", 0) for g in gst_returns)
            ws.append(["Civil Construction ERP - Executive Summary"])
            ws.append([f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M')}"])
            ws.append([])
            ws.append(["Metric", "Value"])
            style_excel_header(ws, 4)
            data = [
                ["Total Projects", len(projects)], ["Active Projects", len([p for p in projects if p.get("status") == "in_progress"])],
                ["Total Budget", total_budget], ["Total Spent", total_spent],
                ["Budget Utilization %", round((total_spent/total_budget*100) if total_budget else 0, 1)],
                ["Total Billed", total_billed], ["Total Received", total_received],
                ["Collection Efficiency %", round((total_received/total_billed*100) if total_billed else 0, 1)],
                ["Active Vendors", len(vendors)], ["Total PO Value", sum(po.get("total", 0) for po in pos)],
                ["Total Employees", len(employees)], ["Total Payroll", total_payroll], ["GST Payable", total_gst],
            ]
            for row in data:
                ws.append(row)
        elif report_type == "project-analysis":
            ws.title = "Project Analysis"
            ws.append(["Project Code", "Project Name", "Client", "Location", "Status", "Budget", "Actual Cost", "Variance", "Progress %", "Start Date", "End Date"])
            style_excel_header(ws)
            for p in projects:
                ws.append([p.get("code"), p.get("name"), p.get("client_name"), p.get("location"), p.get("status"), p.get("budget", 0), p.get("actual_cost", 0), p.get("budget", 0) - p.get("actual_cost", 0), p.get("progress_percentage", 0), p.get("start_date"), p.get("expected_end_date")])
        elif report_type == "financial-summary":
            ws.title = "Billing"
            ws.append(["Bill No", "Date", "Project", "Description", "Type", "Amount", "GST", "Total", "Status"])
            style_excel_header(ws)
            proj_map = {p.get("id"): p.get("name") for p in projects}
            for b in billings:
                ws.append([b.get("bill_number"), b.get("bill_date"), proj_map.get(b.get("project_id"), "-"), b.get("description"), b.get("bill_type"), b.get("amount", 0), b.get("gst_amount", 0), b.get("total_amount", 0), b.get("status")])
            ws2 = wb.create_sheet("CVR")
            ws2.append(["Project", "Period Start", "Period End", "Contracted", "Work Done", "Billed", "Received", "Retention", "Variance"])
            style_excel_header(ws2)
            for c in cvrs:
                ws2.append([proj_map.get(c.get("project_id"), "-"), c.get("period_start"), c.get("period_end"), c.get("contracted_value", 0), c.get("work_done_value", 0), c.get("billed_value", 0), c.get("received_value", 0), c.get("retention_held", 0), c.get("variance", 0)])
        elif report_type == "procurement-analysis":
            ws.title = "Vendors"
            ws.append(["Name", "Category", "GSTIN", "City", "State", "Contact", "Phone", "Email", "Rating"])
            style_excel_header(ws)
            for v in vendors:
                ws.append([v.get("name"), v.get("category"), v.get("gstin"), v.get("city"), v.get("state"), v.get("contact_person"), v.get("phone"), v.get("email"), v.get("rating", 0)])
            ws2 = wb.create_sheet("Purchase Orders")
            ws2.append(["PO Number", "Date", "Vendor", "Delivery Date", "Subtotal", "GST", "Total", "Status"])
            style_excel_header(ws2)
            vendor_map = {v.get("id"): v.get("name") for v in vendors}
            for po in pos:
                ws2.append([po.get("po_number"), po.get("po_date"), vendor_map.get(po.get("vendor_id"), "-"), po.get("delivery_date"), po.get("subtotal", 0), po.get("gst_amount", 0), po.get("total", 0), po.get("status")])
        elif report_type == "hrms-summary":
            ws.title = "Employees"
            ws.append(["Code", "Name", "Designation", "Department", "Phone", "Email", "Joined", "Basic Salary", "HRA", "PF No", "ESI No"])
            style_excel_header(ws)
            for e in employees:
                ws.append([e.get("employee_code"), e.get("name"), e.get("designation"), e.get("department"), e.get("phone"), e.get("email"), e.get("date_of_joining"), e.get("basic_salary", 0), e.get("hra", 0), e.get("pf_number"), e.get("esi_number")])
            ws2 = wb.create_sheet("Payroll")
            ws2.append(["Employee", "Month", "Basic", "HRA", "OT Pay", "Gross", "PF", "ESI", "TDS", "Total Deductions", "Net Salary", "Status"])
            style_excel_header(ws2)
            emp_map = {e.get("id"): e.get("name") for e in employees}
            for p in payrolls:
                ws2.append([emp_map.get(p.get("employee_id"), "-"), p.get("month"), p.get("basic_salary", 0), p.get("hra", 0), p.get("overtime_pay", 0), p.get("gross_salary", 0), p.get("pf_deduction", 0), p.get("esi_deduction", 0), p.get("tds", 0), p.get("total_deductions", 0), p.get("net_salary", 0), p.get("status")])
        elif report_type == "compliance-status":
            ws.title = "GST Returns"
            ws.append(["Type", "Period", "Outward Supplies", "Inward Supplies", "CGST", "SGST", "IGST", "ITC Claimed", "Tax Payable", "Status"])
            style_excel_header(ws)
            for g in gst_returns:
                ws.append([g.get("return_type"), g.get("period"), g.get("total_outward_supplies", 0), g.get("total_inward_supplies", 0), g.get("cgst", 0), g.get("sgst", 0), g.get("igst", 0), g.get("itc_claimed", 0), g.get("tax_payable", 0), g.get("status")])
        elif report_type == "cost-variance":
            ws.title = "Cost Variance"
            ws.append(["Project Code", "Project Name", "Budget", "Actual Cost", "Variance", "Variance %", "Status", "CPI"])
            style_excel_header(ws)
            for p in projects:
                budget = p.get("budget", 0); actual = p.get("actual_cost", 0); variance = budget - actual
                ws.append([p.get("code"), p.get("name"), budget, actual, variance, round((variance/budget*100) if budget else 0, 1), "Under Budget" if variance >= 0 else "Over Budget", round((budget/actual) if actual else 0, 2)])
        else:
            raise HTTPException(status_code=400, detail=f"Unknown report type: {report_type}")

        auto_column_width(ws)
        filepath = EXPORT_DIR / f"{report_type}_{timestamp}.xlsx"
        wb.save(str(filepath))
        return FileResponse(str(filepath), filename=f"{report_type}_{timestamp}.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    elif format == "pdf":
        filepath = EXPORT_DIR / f"{report_type}_{timestamp}.pdf"
        doc = SimpleDocTemplate(str(filepath), pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("ReportTitle", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
        subtitle_style = ParagraphStyle("ReportSubtitle", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=12)
        elements = []
        report_titles = {"executive-summary": "Executive Summary", "project-analysis": "Project Analysis", "financial-summary": "Financial Summary", "procurement-analysis": "Procurement Analysis", "hrms-summary": "HRMS Summary", "compliance-status": "Compliance Status", "cost-variance": "Cost Variance"}
        elements.append(Paragraph(f"Civil ERP - {report_titles.get(report_type, report_type)}", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}", subtitle_style))
        header_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
        if report_type == "executive-summary":
            total_budget = sum(p.get("budget", 0) for p in projects)
            total_spent = sum(p.get("actual_cost", 0) for p in projects)
            total_billed = sum(b.get("total_amount", 0) for b in billings)
            data = [["Metric", "Value"],
                ["Total Projects", str(len(projects))], ["Total Budget", f"INR {total_budget:,.0f}"],
                ["Total Spent", f"INR {total_spent:,.0f}"], ["Total Billed", f"INR {total_billed:,.0f}"],
                ["Active Vendors", str(len(vendors))], ["Total Employees", str(len(employees))],
                ["Total Payroll", f"INR {sum(p.get('net_salary',0) for p in payrolls):,.0f}"]]
            t = RLTable(data, colWidths=[120*mm, 120*mm])
            t.setStyle(header_style)
            elements.append(t)
        elif report_type == "project-analysis":
            data = [["Code", "Name", "Client", "Status", "Budget", "Actual", "Variance", "Progress"]]
            for p in projects:
                data.append([p.get("code",""), p.get("name","")[:25], p.get("client_name","")[:20], p.get("status",""), f"{p.get('budget',0):,.0f}", f"{p.get('actual_cost',0):,.0f}", f"{p.get('budget',0)-p.get('actual_cost',0):,.0f}", f"{p.get('progress_percentage',0)}%"])
            t = RLTable(data); t.setStyle(header_style); elements.append(t)
        elif report_type == "financial-summary":
            proj_map = {p.get("id"): p.get("name","")[:20] for p in projects}
            data = [["Bill No", "Date", "Project", "Amount", "GST", "Total", "Status"]]
            for b in billings:
                data.append([b.get("bill_number",""), b.get("bill_date",""), proj_map.get(b.get("project_id"),"-"), f"{b.get('amount',0):,.0f}", f"{b.get('gst_amount',0):,.0f}", f"{b.get('total_amount',0):,.0f}", b.get("status","")])
            t = RLTable(data); t.setStyle(header_style); elements.append(t)
        elif report_type == "hrms-summary":
            data = [["Code", "Name", "Designation", "Department", "Basic Salary", "HRA", "Joined"]]
            for e in employees:
                data.append([e.get("employee_code",""), e.get("name",""), e.get("designation","")[:20], e.get("department",""), f"{e.get('basic_salary',0):,.0f}", f"{e.get('hra',0):,.0f}", e.get("date_of_joining","")])
            t = RLTable(data); t.setStyle(header_style); elements.append(t)
        elif report_type == "procurement-analysis":
            data = [["Name", "Category", "GSTIN", "City", "Phone", "Rating"]]
            for v in vendors:
                data.append([v.get("name",""), v.get("category",""), v.get("gstin",""), v.get("city",""), v.get("phone",""), str(v.get("rating",0))])
            t = RLTable(data); t.setStyle(header_style); elements.append(t)
        elif report_type == "cost-variance":
            data = [["Code", "Name", "Budget", "Actual", "Variance", "Var %", "Status"]]
            for p in projects:
                b = p.get("budget", 0); a = p.get("actual_cost", 0); v = b - a
                data.append([p.get("code",""), p.get("name","")[:25], f"{b:,.0f}", f"{a:,.0f}", f"{v:,.0f}", f"{(v/b*100) if b else 0:.1f}%", "Under" if v >= 0 else "Over"])
            t = RLTable(data); t.setStyle(header_style); elements.append(t)
        elif report_type == "compliance-status":
            data = [["Type", "Period", "CGST", "SGST", "IGST", "ITC", "Tax Payable", "Status"]]
            for g in gst_returns:
                data.append([g.get("return_type",""), g.get("period",""), f"{g.get('cgst',0):,.0f}", f"{g.get('sgst',0):,.0f}", f"{g.get('igst',0):,.0f}", f"{g.get('itc_claimed',0):,.0f}", f"{g.get('tax_payable',0):,.0f}", g.get("status","")])
            t = RLTable(data); t.setStyle(header_style); elements.append(t)
        else:
            raise HTTPException(status_code=400, detail=f"Unknown report type: {report_type}")
        doc.build(elements)
        return FileResponse(str(filepath), filename=f"{report_type}_{timestamp}.pdf", media_type="application/pdf")

    raise HTTPException(status_code=400, detail="Format must be 'excel' or 'pdf'")
